"""End-to-end flow: competition -> upload -> categories -> round -> export."""
import io

from openpyxl import Workbook, load_workbook

from tests.conftest import register


def _sample_xlsx():
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Kadeci")
    ws.append(["Name and Surname", "Year of Birth", "Weight category", "Team"])
    for i in range(1, 8):  # 7 wrestlers in one weight class
        ws.append([f"Zawodnik {i}", 2011, "45 kg", "UKS Test"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _logged_in(client):
    register(client, "tester", email="tester@x.pl")
    return client


def test_full_flow(client):
    _logged_in(client)
    # create competition
    r = client.post(
        "/competitions",
        data={"name": "Test Cup", "event_date": "2026-07-01", "birth_years": "2011"},
        follow_redirects=False,
    )
    assert r.status_code == 303
    comp_url = r.headers["location"]
    comp_id = int(comp_url.rstrip("/").split("/")[-1])

    # upload documents
    r = client.post(
        f"/competitions/{comp_id}/documents",
        files={"files": ("kadeci.xlsx", _sample_xlsx(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    assert "Wczytano uczestników" in r.text

    # competition page lists the age category
    r = client.get(f"/competitions/{comp_id}")
    assert "Kadeci" in r.text

    # find the age category id and open it
    page = client.get(f"/competitions/{comp_id}").text
    import re
    age_id = int(re.search(r"/age-categories/(\d+)", page).group(1))
    r = client.get(f"/age-categories/{age_id}")
    assert "45 kg" in r.text
    weight_id = int(re.search(r"/weight-categories/(\d+)", r.text).group(1))

    # weight category page shows 7 participants
    r = client.get(f"/weight-categories/{weight_id}")
    assert "Zawodnik 1" in r.text

    # shuffle
    r = client.post(f"/weight-categories/{weight_id}/shuffle")
    assert r.status_code == 200
    assert "Number" in r.text

    # create round 1 (auto-generated bracket)
    r = client.post(f"/weight-categories/{weight_id}/rounds", follow_redirects=False)
    assert r.status_code == 303
    page = client.get(f"/weight-categories/{weight_id}").text
    round_id = int(re.search(r"/rounds/(\d+)", page).group(1))

    # open the round form -> 7 wrestlers => split into two groups + FINAŁ
    r = client.get(f"/rounds/{round_id}")
    assert r.status_code == 200
    assert "FINAŁ" in r.text

    # downloads
    r = client.get(f"/rounds/{round_id}/excel")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.active.title == "Protokół"

    r = client.get(f"/rounds/{round_id}/pdf")
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"

    r = client.get(f"/rounds/{round_id}/pdf-judgement")
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"


def test_permissions_block_stranger(client):
    # owner creates a competition
    _logged_in(client)
    r = client.post("/competitions", data={"name": "Private"}, follow_redirects=False)
    comp_id = int(r.headers["location"].rstrip("/").split("/")[-1])

    # a different user has no access
    other = client
    other.get("/logout")
    register(other, "intruder", email="intruder@x.pl")
    r = other.get(f"/competitions/{comp_id}", follow_redirects=False)
    assert r.status_code == 403


def test_share_grants_access(client):
    _logged_in(client)
    r = client.post("/competitions", data={"name": "Shared"}, follow_redirects=False)
    comp_id = int(r.headers["location"].rstrip("/").split("/")[-1])

    # create the target user in a separate client
    from fastapi.testclient import TestClient
    from app.main import app
    friend = TestClient(app)
    register(friend, "friend", email="friend@x.pl")

    # owner shares (read) with friend
    r = client.post(
        f"/competitions/{comp_id}/share",
        data={"target": "friend", "level": "read"},
    )
    assert "Udostępniono" in r.text

    # friend can now read
    r = friend.get(f"/competitions/{comp_id}", follow_redirects=False)
    assert r.status_code == 200
