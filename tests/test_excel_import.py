"""Tests for Excel parsing and import."""
import io

from openpyxl import Workbook

from app.services.excel_import import parse_workbook


def _wb_bytes():
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Kadeci")
    ws.append(["Name and Surname", "Year of Birth", "Weight category", "Team", "Pas"])
    ws.append(["Jan Nowak", 2011, "45 kg", "UKS Warszawa", "żółty"])
    ws.append(["Piotr Kowalski", 2012, "45 kg", "KS Kraków", None])
    ws.append([None, None, None, None, None])  # blank row ignored
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_workbook_columns_and_other():
    sheets = parse_workbook(_wb_bytes())
    assert len(sheets) == 1
    sheet = sheets[0]
    assert sheet.age_category == "Kadeci"
    assert len(sheet.participants) == 2
    p = sheet.participants[0]
    assert p.name == "Jan Nowak"
    assert p.year == 2011
    assert p.weight == "45 kg"
    assert p.team == "UKS Warszawa"
    # extra column folded into "other"
    assert "Pas" in p.other


def test_parse_polish_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Młodzicy"
    ws.append(["Nazwisko i imię", "Rok", "Kategoria wagowa", "Klub"])
    ws.append(["Maja Wójcik", 2014, "30 kg", "AZS Wrocław"])
    buf = io.BytesIO()
    wb.save(buf)
    sheets = parse_workbook(buf.getvalue())
    assert sheets[0].participants[0].name == "Maja Wójcik"
    assert sheets[0].participants[0].weight == "30 kg"
