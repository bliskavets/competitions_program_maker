"""Tests for the Excel / PDF exporters."""
import io

from openpyxl import load_workbook

from app.services.brackets import build_empty_bracket, build_initial_bracket
from app.services.exports_excel import render_bracket_xlsx
from app.services.exports_pdf import render_bracket_pdf


def _people(n):
    return [{"name": f"Z{i}", "year": 2010, "team": "K"} for i in range(1, n + 1)]


def test_excel_round_robin_has_judging_column():
    bracket = build_initial_bracket(_people(4))
    plain = render_bracket_xlsx(bracket, title="T", judging=False)
    judge = render_bracket_xlsx(bracket, title="T", judging=True)
    wb = load_workbook(io.BytesIO(judge))
    judge_cells = {c.value for row in wb.active.iter_rows() for c in row}
    # the "Win" column only appears in the judging variant
    assert "Win" in judge_cells
    wb2 = load_workbook(io.BytesIO(plain))
    plain_cells = {c.value for row in wb2.active.iter_rows() for c in row}
    assert "Win" not in plain_cells


def test_excel_is_a4_fit_one_page():
    bracket = build_initial_bracket(_people(10))
    data = render_bracket_xlsx(bracket, title="A4")
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 1


def test_pdf_renders_for_all_sizes():
    for n in (1, 3, 5, 6, 10, 16, 33):
        bracket = build_initial_bracket(_people(n))
        pdf = render_bracket_pdf(bracket, title=f"n={n}")
        assert pdf[:4] == b"%PDF"
        pdf_j = render_bracket_pdf(bracket, title=f"n={n}", judging=True)
        assert pdf_j[:4] == b"%PDF"


def test_empty_bracket_export():
    bracket = build_empty_bracket(12)
    assert render_bracket_xlsx(bracket, title="empty")[:2] == b"PK"  # xlsx = zip
    assert render_bracket_pdf(bracket, title="empty")[:4] == b"%PDF"
