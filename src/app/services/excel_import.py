"""Import participants from Excel workbooks.

Each workbook has one sheet per age category. Required columns (case- and
whitespace-insensitive):

    Name and Surname | Year of Birth | Weight category | Team

Any further columns are optional and merged into the participant's
"Other Information".
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import AgeCategory, Competition, Participant, WeightCategory

REQUIRED = {
    "name": ("name and surname", "name", "nazwisko i imię", "imię i nazwisko"),
    "year": ("year of birth", "rok", "rok urodzenia"),
    "weight": ("weight category", "weight", "kategoria wagowa", "waga"),
    "team": ("team", "klub", "drużyna"),
}


def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _match_columns(header: list[Any]) -> dict[str, int]:
    """Map logical field -> column index based on the header row."""
    mapping: dict[str, int] = {}
    norm_header = [_norm(h) for h in header]
    for field_name, aliases in REQUIRED.items():
        for idx, h in enumerate(norm_header):
            if h in aliases:
                mapping[field_name] = idx
                break
    return mapping


@dataclass
class ParsedParticipant:
    name: str
    year: int | None
    weight: str
    team: str | None
    other: dict[str, Any] = field(default_factory=dict)


@dataclass
class ParsedSheet:
    age_category: str
    participants: list[ParsedParticipant]


def parse_workbook(data: bytes) -> list[ParsedSheet]:
    """Parse raw .xlsx bytes into structured sheets without touching the DB."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets: list[ParsedSheet] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = list(rows[0])
        cols = _match_columns(header)
        if "name" not in cols or "weight" not in cols:
            # Not a participant sheet (e.g. metadata) — skip.
            continue
        participants: list[ParsedParticipant] = []
        for raw in rows[1:]:
            row = list(raw)
            name = _cell(row, cols.get("name"))
            if not name or not str(name).strip():
                continue
            weight = _cell(row, cols.get("weight"))
            other = {}
            for i, value in enumerate(row):
                if i not in cols.values() and value not in (None, ""):
                    key = str(header[i]) if i < len(header) and header[i] else f"col{i}"
                    other[key] = value
            participants.append(
                ParsedParticipant(
                    name=str(name).strip(),
                    year=_to_int(_cell(row, cols.get("year"))),
                    weight=str(weight).strip() if weight is not None else "",
                    team=_opt_str(_cell(row, cols.get("team"))),
                    other=other,
                )
            )
        if participants:
            sheets.append(ParsedSheet(age_category=ws.title, participants=participants))
    wb.close()
    return sheets


def import_workbook(db: Session, competition: Competition, data: bytes) -> int:
    """Parse and persist a workbook into the competition. Returns count added."""
    sheets = parse_workbook(data)
    added = 0
    for sheet in sheets:
        age = _get_or_create_age_category(db, competition, sheet.age_category)
        for p in sheet.participants:
            weight = _get_or_create_weight_category(db, age, p.weight)
            other_text = (
                "; ".join(f"{k}: {v}" for k, v in p.other.items()) or None
            )
            db.add(
                Participant(
                    weight_category_id=weight.id,
                    name=p.name,
                    birth_year=p.year,
                    team=p.team,
                    other_info=other_text,
                )
            )
            added += 1
        db.flush()
    db.commit()
    return added


def _get_or_create_age_category(
    db: Session, competition: Competition, title: str
) -> AgeCategory:
    title = title.strip()
    for ac in competition.age_categories:
        if (ac.name or "").strip().lower() == title.lower():
            return ac
    ac = AgeCategory(competition_id=competition.id, name=title)
    db.add(ac)
    db.flush()
    db.refresh(competition)
    return ac


def _get_or_create_weight_category(
    db: Session, age: AgeCategory, weight_label: str
) -> WeightCategory:
    weight_value = _to_float(weight_label)
    label = (weight_label or "").strip()
    for wc in age.weight_categories:
        if (wc.name or "").strip().lower() == label.lower():
            return wc
    wc = WeightCategory(age_category_id=age.id, name=label, weight=weight_value)
    db.add(wc)
    db.flush()
    db.refresh(age)
    return wc


def _cell(row: list[Any], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _opt_str(v: Any) -> str | None:
    if v is None or str(v).strip() == "":
        return None
    return str(v).strip()


def _to_int(v: Any) -> int | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(v))
    if not m:
        return None
    try:
        return float(m.group().replace(",", "."))
    except ValueError:
        return None
