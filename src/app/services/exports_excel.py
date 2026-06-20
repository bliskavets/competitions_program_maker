"""Render tournament sheets to A4-printable .xlsx (openpyxl).

A round-robin group is rendered as the classic "krzyżówka" table:

    Lp | Nazwisko i imię | Rok | Klub | Rundy 1..k | suma PKT | M-ce [| Win]

For the *judging* variant an extra ``Win`` column is appended so the referee
can tick the winner of each bout. Page setup is fixed to A4 fit-to-one-page.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="E8E8E8")
REST_FILL = PatternFill("solid", fgColor="D9F2F2")  # the "wl" / bye cells
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BOLD = Font(bold=True)


def render_bracket_xlsx(
    bracket: dict[str, Any],
    *,
    title: str = "",
    judging: bool = False,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Protokół"
    _setup_a4(ws)

    row = 1
    if title:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=14)
        row += 2

    groups = bracket.get("groups") or []
    if not groups and bracket.get("type") == "single_elim":
        groups = [bracket]

    for group in groups:
        gname = group.get("name")
        if gname:
            ws.cell(row=row, column=1, value=f"Grupa {gname}").font = BOLD
            row += 1
        if group.get("type") == "single_elim" or "rounds" in group and "schedule" not in group:
            row = _render_single_elim(ws, group, row, judging)
        else:
            row = _render_round_robin(ws, group, row, judging)
        row += 1

    if bracket.get("final"):
        row = _render_final(ws, bracket["final"], row)

    _footer(ws, row + 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_round_robin(
    ws: Worksheet, group: dict[str, Any], start: int, judging: bool
) -> int:
    num_rounds = max(int(group.get("num_rounds", 0)), 1)
    headers = ["Lp.", "Nazwisko i imię", "Rok", "Klub"]
    headers += [str(i) for i in range(1, num_rounds + 1)]
    headers += ["suma PKT", "M-ce"]
    if judging:
        headers.append("Win")

    r = start
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    r += 1

    for row_data in group.get("rows", []):
        # round columns show the opponent's position (or "wl" for the bye)
        cells = row_data.get("cells") or [None] * num_rounds
        round_cells = ["wl" if o is None else o for o in cells]
        values = [
            row_data.get("lp"),
            row_data.get("name"),
            row_data.get("year"),
            row_data.get("team"),
        ]
        values += round_cells
        values += [None, None]  # suma PKT, M-ce
        if judging:
            values.append(None)
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = LEFT if c == 2 else CENTER
            # grey out the bye ("wl") cell
            if 5 <= c <= 4 + num_rounds and v == "wl":
                cell.fill = REST_FILL
        r += 1
    return r


def _render_single_elim(
    ws: Worksheet, group: dict[str, Any], start: int, judging: bool
) -> int:
    r = start
    rounds = group.get("rounds", [])
    # Header row listing round columns
    ws.cell(row=r, column=1, value="Runda").font = BOLD
    ws.cell(row=r, column=1).border = BORDER
    for ri, rnd in enumerate(rounds, start=1):
        cell = ws.cell(row=r, column=1 + ri, value=f"R{rnd.get('round', ri)}")
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    r += 1

    rows = group.get("rows", [])
    if rows:
        for row_data in rows:
            ws.cell(
                row=r,
                column=1,
                value=f"{row_data.get('lp')}. {row_data.get('name','')}",
            ).border = BORDER
            r += 1
        r += 1

    # Show first-round pairings and byes ("wolny")
    first = rounds[0]["matches"] if rounds else []
    ws.cell(row=r, column=1, value="Pary 1. rundy:").font = BOLD
    r += 1
    for i, m in enumerate(first, start=1):
        top = m.get("top")
        bottom = m.get("bottom")
        if bottom is None:
            label = f"{i}. nr {top} — wolny"
            cell = ws.cell(row=r, column=1, value=label)
            cell.fill = PatternFill("solid", fgColor="EFEFEF")
            cell.font = Font(color="888888")
        else:
            cell = ws.cell(row=r, column=1, value=f"{i}. nr {top} — nr {bottom}")
        cell.border = BORDER
        if judging:
            ws.cell(row=r, column=2, value="Win:").border = BORDER
        r += 1
    return r


def _render_final(ws: Worksheet, final: dict[str, Any], start: int) -> int:
    r = start + 1
    ws.cell(row=r, column=1, value="FINAŁ").font = Font(bold=True, size=12)
    r += 1
    for a, b in final.get("pairs", []):
        ws.cell(row=r, column=1, value=a).border = BORDER
        ws.cell(row=r, column=2, value="—").alignment = CENTER
        ws.cell(row=r, column=3, value=b).border = BORDER
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="M-ce").font = BOLD
    ws.cell(row=r, column=2, value="Nazwisko i imię").font = BOLD
    r += 1
    for place in final.get("places", []):
        ws.cell(row=r, column=1, value=place).border = BORDER
        ws.cell(row=r, column=2, value=None).border = BORDER
        r += 1
    return r


def _footer(ws: Worksheet, row: int) -> None:
    ws.cell(row=row, column=1, value="Kierownik List").alignment = CENTER
    ws.cell(row=row, column=5, value="Sędzia Główny").alignment = CENTER


def _setup_a4(ws: Worksheet) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.3
    # Reasonable default widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 18
    for col in range(5, 20):
        ws.column_dimensions[get_column_letter(col)].width = 6
